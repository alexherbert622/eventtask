import pandas as pd
from openpyxl import load_workbook


class Datareadfile:

    df = pd.read_excel('G:\\Thirdeye\\eventtask.xlsx', engine='openpyxl')
    url = df.iloc[0][5]
    username = df.iloc[1][5]
    password = df.iloc[2][5]


def datawritepass():
    filepath = 'G:\\Thirdeye\\eventtask.xlsx'
    ws = load_workbook(filepath)
    wb = ws.worksheets[0]
    wb["G2"] = "url loaded successfully"
    wb["H2"] = "pass"
    wb["G3"] = "Username entered correctly"
    wb["H3"] = "pass"
    wb["G4"] = "password entered correctly"
    wb["H4"] = "pass"
    wb["G5"] = "logged in successfully"
    wb["H5"] = "pass"
    ws.save(filepath)


def datawritefail():

    filepath = 'G:\\Thirdeye\\eventtask.xlsx'
    ws = load_workbook(filepath)
    wb = ws.worksheets[0]
    wb["G2"] = "url loaded unsuccessfully"
    wb["H2"] = "fail"
    wb["G3"] = "Username entered incorrectly"
    wb["H3"] = "fail"
    wb["G4"] = "password entered incorrectly"
    wb["H4"] = "fail"
    wb["G5"] = "logged in unsuccessfully"
    wb["H5"] = "fail"
    ws.save(filepath)